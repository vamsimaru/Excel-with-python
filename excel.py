from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
wb = load_workbook('/Downloads/Lists.xlsx')
new_sheet = wb.create_sheet('List 3', 2)
df_1 = pd.read_excel('/Downloads/Lists.xlsx', sheet_name='List 1')
df_1 = pd.read_excel('/Downloads/Lists.xlsx', sheet_name='List 2')
